import openpyxl
import zipfile
import os


# 解决解压乱码问题
def support_gbk(zip_file: zipfile.ZipFile):
    name_to_info = zip_file.NameToInfo
    # copy map first
    for name, info in name_to_info.copy().items():
        real_name = name.encode('cp437').decode('gbk')
        if real_name != name:
            info.filename = real_name
            del name_to_info[name]
            name_to_info[real_name] = info
    return zip_file


if not os.path.exists("E:/销售单"):
    os.mkdir("E:/销售单")
with support_gbk(zipfile.ZipFile("销售单.zip", "r")) as zip_ref:
    zip_ref.extractall("E:/销售单")


def sort_ren(wb_s, save_path, ws):
    sheet1 = wb_s[wb_s.sheetnames[0]]
    par01 = set()
    for item in sheet1["F3:G23"]:  # item表示每一行的单元格元组
        par01.add(item[0].value)
    # 初始化字典
    dict01 = {}
    for i in par01:
        dict01[i] = 0
    # 统计每个人的销售量
    for item in sheet1["E3:G23"]:
        dict01[item[1].value] += int(item[2].value) * int(item[0].value)
    k = []
    for key in dict01:
        num = [key, dict01[key]]
        k.append(num)
    k = sorted(k, key=(lambda x: x[1]), reverse=True)
    # k.sort(key=(lambda x: x[1]))
    k.insert(0, ["姓名", "销售额"])
    for i in k:
        ws.append(i)
    # 保存 Excel 文件
    wb.save(save_path)


def sort_che(wb_s, save_path, ws):
    sheet1 = wb_s[wb_s.sheetnames[0]]
    par01 = set()
    for item in sheet1["C3:G23"]:  # item表示每一行的单元格元组
        par01.add(item[0].value)
    # 初始化字典
    dict01 = {}
    for i in par01:
        dict01[i] = 0
    dict02 = {}
    for i in par01:
        dict02[i] = 0
    # 统计每个人的销售量
    for item in sheet1["C3:G23"]:
        # 品牌销售额
        dict01[item[0].value] += int(item[2].value) * int(item[4].value)
        # 品牌销售数量
        dict02[item[0].value] += int(item[2].value)
    # 字典转换为列表
    k1 = []
    for key in dict01:
        num = [key, dict01[key]]
        k1.append(num)

    k2 = []
    for key in dict02:
        num = [key, dict02[key]]
        k2.append(num)

    for i in k1:
        for j in k2:
            if i[0] == j[0]:
                i.append(j[1])

    k1 = sorted(k1, key=(lambda x: x[1]), reverse=True)
    # k.sort(key=(lambda x: x[1]))
    # k2 = sorted(k2, key=(lambda x: x[1]), reverse=True)

    k1.insert(0, ["品牌", "销售额", "销售数量"])
    # k2.insert(0, ["品牌", "销售数量"])

    for i in k1:
        ws.append(i)
    # 保存 Excel 文件
    wb.save(save_path)


# 创建排序后的表格
dir01 = "E:/销售单/销售单"
ls01 = os.listdir(dir01)
wb_s1 = openpyxl.load_workbook(os.path.join(dir01, ls01[0]))
wb_s2 = openpyxl.load_workbook(os.path.join(dir01, ls01[1]))
wb_s3 = openpyxl.load_workbook(os.path.join(dir01, ls01[2]))

save_path = "E:/销售单/汽车销售统计.xlsx"

# 创建一个 workbook
wb = openpyxl.Workbook()
# 获取被激活的 worksheet
ws1 = wb.active
ws1.title = "4月销售排行榜"
ws2 = wb.create_sheet("5月销售排行榜", 1)
ws3 = wb.create_sheet("6月销售排行榜", 2)

ws01 = wb.create_sheet("4月品牌销售排行榜", 3)
ws02 = wb.create_sheet("5月品牌销售排行榜", 4)
ws03 = wb.create_sheet("6月品牌销售排行榜", 5)

sort_ren(wb_s1, save_path, ws1)
sort_ren(wb_s2, save_path, ws2)
sort_ren(wb_s3, save_path, ws3)

sort_che(wb_s1, save_path, ws01)
sort_che(wb_s2, save_path, ws02)
sort_che(wb_s3, save_path, ws03)

